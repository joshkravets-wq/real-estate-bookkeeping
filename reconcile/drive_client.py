"""
Drive client for the reconciliation engine.

Wraps Google Drive API auth + file fetching. Exposes simple methods
for the two file types we care about:
  - .xlsx files: fetched as bytes, parsed with openpyxl
  - Google Sheets (.gsheet): exported as xlsx bytes, parsed the same way

Auth lives in credentials/token.json (gitignored). If the token is missing
or expired, the script tells the user to re-run authorize.py.

Usage:
    from reconcile.drive_client import DriveClient
    client = DriveClient()
    workbook = client.fetch_spreadsheet(file_id)  # openpyxl.Workbook
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Optional

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook


SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/spreadsheets.readonly",
]

REPO_ROOT = Path(__file__).parent.parent
CREDS_DIR = REPO_ROOT / "credentials"
TOKEN_FILE = CREDS_DIR / "token.json"

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
GOOGLE_SHEET_MIME = "application/vnd.google-apps.spreadsheet"


class DriveAuthError(Exception):
    """Raised when token is missing or expired."""
    pass


CACHE_DIR = CREDS_DIR / "cache"


class DriveClient:
    def __init__(self, token_file: Optional[Path] = None):
        self.token_file = token_file or TOKEN_FILE
        self.offline = False
        if not self.token_file.exists():
            # Offline-cache mode: if credentials/cache/<file_id>.xlsx files
            # exist (pre-downloaded by another Drive-authorized channel),
            # serve spreadsheets from there instead of the live API.
            if CACHE_DIR.is_dir() and any(CACHE_DIR.glob("*.xlsx")):
                self.offline = True
                self.creds = None
                self.service = None
                return
            raise DriveAuthError(
                f"Missing {self.token_file}. Run credentials/authorize.py first."
            )
        self.creds = Credentials.from_authorized_user_file(str(self.token_file), SCOPES)

        # Refresh if expired (silent — uses refresh token)
        if self.creds.expired and self.creds.refresh_token:
            self.creds.refresh(Request())
            with self.token_file.open("w") as f:
                f.write(self.creds.to_json())

        self.service = build("drive", "v3", credentials=self.creds, cache_discovery=False)

    def get_metadata(self, file_id: str) -> dict:
        """Fetch basic metadata for a file. Useful to check mime type before download."""
        if self.offline:
            path = CACHE_DIR / f"{file_id}.xlsx"
            if not path.exists():
                raise FileNotFoundError(
                    f"Offline cache miss: {path}. Pre-download this sheet into credentials/cache/."
                )
            return {
                "id": file_id,
                "name": path.name,
                "mimeType": XLSX_MIME,
                "size": str(path.stat().st_size),
                "modifiedTime": None,
            }
        return self.service.files().get(
            fileId=file_id,
            fields="id,name,mimeType,size,modifiedTime",
        ).execute()

    def _download_xlsx_bytes(self, file_id: str) -> bytes:
        """Download an xlsx file's raw bytes."""
        request = self.service.files().get_media(fileId=file_id)
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        return buf.getvalue()

    def _export_sheet_as_xlsx_bytes(self, file_id: str) -> bytes:
        """Export a Google Sheet as xlsx bytes."""
        request = self.service.files().export_media(
            fileId=file_id, mimeType=XLSX_MIME
        )
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        return buf.getvalue()

    def fetch_spreadsheet(self, file_id: str):
        """Return an openpyxl Workbook for any spreadsheet (xlsx or Google Sheet).

        Caller doesn't need to know the type. We check metadata, download
        in the correct mode, and parse.
        """
        if self.offline:
            path = CACHE_DIR / f"{file_id}.xlsx"
            if not path.exists():
                raise FileNotFoundError(
                    f"Offline cache miss: {path}. Pre-download this sheet into credentials/cache/."
                )
            return load_workbook(str(path), data_only=True)

        meta = self.get_metadata(file_id)
        mime = meta["mimeType"]

        if mime == XLSX_MIME:
            data = self._download_xlsx_bytes(file_id)
        elif mime == GOOGLE_SHEET_MIME:
            data = self._export_sheet_as_xlsx_bytes(file_id)
        else:
            raise ValueError(
                f"File {meta['name']} has mime type {mime}; not a spreadsheet"
            )

        return load_workbook(io.BytesIO(data), data_only=True)


# ---------------------------------------------------------------------------
# CLI smoke test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    client = DriveClient()

    # Test 1: xlsx file (Bank Credits & Debits)
    print("Test 1: xlsx file")
    wb = client.fetch_spreadsheet("1l5ujV9j5EKd32_cHHHToCzDO-F1nXDh2")
    sheet = wb.active
    print(f"  Sheet name: {sheet.title}")
    print(f"  Dimensions: {sheet.dimensions}")
    print(f"  Row 1: {[c.value for c in sheet[1]]}")
    print(f"  Row 2: {[c.value for c in sheet[2]]}")
    print()

    # Test 2: Google Sheet (5461 W Berks)
    print("Test 2: Google Sheet")
    wb = client.fetch_spreadsheet("1tp1IX4hqlmPXKhYGDkuKa91rCNx5xklyoMW6TZpTSbs")
    print(f"  Sheet names: {wb.sheetnames}")
    sheet = wb.active
    print(f"  Active sheet: {sheet.title}")
    print(f"  Dimensions: {sheet.dimensions}")
    print(f"  Row 1: {[c.value for c in sheet[1]]}")
    print(f"  Row 2: {[c.value for c in sheet[2]]}")
